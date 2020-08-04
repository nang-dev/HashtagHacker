import xlrd # pip install --user xlrd
import random
import openpyxl # pip install --user openpyxl
  
excel_file = "main.xlsx"
# To open Workbook 
wb = xlrd.open_workbook(excel_file) 
sheet = wb.sheet_by_index(0)

hashtags = []
# For row 0 and column 0 
row = 0
col = 0

while(col < sheet.ncols and ("group" in (sheet.cell_value(0, col)).lower())):
    row = 1
    while(row < sheet.nrows and (sheet.cell_value(row, col)) != ""):
        ht = sheet.cell_value(row, col)
        hashtags.append(ht.strip("#"))
        row += 1
    col += 1

wb = openpyxl.load_workbook(excel_file)
try:
    wb["Sheet1"]
except:
    wb.create_sheet("Sheet1") 
sheet = wb["Sheet1"]

col = 13
for i in range(1,41):
    sampled_list = random.sample(hashtags, 30)
    result_file = open("results/result"+str(i)+".txt", "w")
    sheet.cell(row=1, column=col).value = "Results " + str(i)
    row = 2
    for ht in sampled_list:
        sheet.cell(row=row, column=col).value = ht
        result_file.write("#" + ht + " ")
        row += 1
    col += 1

wb.save(excel_file)

print("Hashtag Hacker Worked!")
print("SUCCESS: Generated text files in 'Results' folder for easy copy-paste!")
print("SUCCESS: Updated excel sheet with results (may need to close and reopen excel page)")